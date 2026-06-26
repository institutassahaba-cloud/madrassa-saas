#!/usr/bin/env python3
"""Extrait l'emploi du temps (onglet 'EDT') d'un classeur prof .xlsx.

Grille hebdo : col A = créneau horaire, cols B..H = Lundi..Dimanche,
chaque cellule = nom de l'élève. Col K et au-delà = listes annexes (ignorées).

Sortie JSON : [{day, start, end, name, raw_time}] sur stdout.
day : 1=Lundi .. 6=Samedi, 0=Dimanche (convention TimeSlot.dayOfWeek).

Usage : python3 _extract_edt.py "/chemin/Fichier prof.xlsx" > edt-PRxxx.json
"""
import sys, json, re
import openpyxl

# colonne B..H -> dayOfWeek (0=Dim..6=Sam)
COL_TO_DAY = {2: 1, 3: 2, 4: 3, 5: 4, 6: 5, 7: 6, 8: 0}  # Lun..Dim


def parse_one_time(tok: str):
    """'8h30' -> (8,30) ; '8h' -> (8,0) ; '19h,30' -> (19,30) ; '20,30' -> (20,30) ; '22,00' -> (22,0)."""
    nums = re.findall(r"\d+", tok)
    if not nums:
        return None
    h = int(nums[0])
    m = int(nums[1]) if len(nums) > 1 else 0
    if h > 23 or m > 59:
        return None
    return h, m


def parse_range(raw: str):
    """'8h-8h30' -> ('08:00','08:30'). Gère séparateurs - et _ entre début/fin."""
    if not raw:
        return None
    s = str(raw).strip()
    # séparateur de plage : '-' ou '_' (mais pas la virgule décimale)
    parts = re.split(r"[-_]", s)
    if len(parts) < 2:
        return None
    a = parse_one_time(parts[0])
    b = parse_one_time(parts[-1])
    if not a or not b:
        return None
    return f"{a[0]:02d}:{a[1]:02d}", f"{b[0]:02d}:{b[1]:02d}"


def find_edt_tab(names):
    """Nom d'onglet EDT variable selon les profs : 'EDT' ou 'EMPLOI DU TEMPS'."""
    for n in names:
        if n.strip().upper() == "EDT":
            return n
    for n in names:
        u = n.strip().upper()
        if "EDT" in u or "EMPLOI" in u:
            return n
    return None


def main(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    tab = find_edt_tab(wb.sheetnames)
    if not tab:
        print(json.dumps({"error": "pas d'onglet EDT", "sheets": wb.sheetnames[:5]}))
        return
    ws = wb[tab]
    slots = []
    for r in range(1, ws.max_row + 1):
        rng = parse_range(ws.cell(row=r, column=1).value)
        if not rng:
            continue
        start, end = rng
        for col, day in COL_TO_DAY.items():
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            name = str(v).strip()
            if not name:
                continue
            slots.append({
                "day": day, "start": start, "end": end,
                "name": name, "raw_time": str(ws.cell(row=r, column=1).value).strip(),
            })
    wb.close()
    print(json.dumps(slots, ensure_ascii=False, indent=1))


if __name__ == "__main__":
    main(sys.argv[1])
